import time
start_time = time.time()

import os

import tkinter as tk
from tkinter import filedialog
    
# Hide the root window
root = tk.Tk()
root.withdraw()

# Ask the user to select a folder
base_dir = filedialog.askdirectory(title="Select Base Directory")

print("Selected folder:", base_dir)


os.chdir(base_dir)

for root, dirs, files in os.walk(base_dir):
    if root not in os.sys.path:
        os.sys.path.append(root)


for path in os.sys.path:
    print(path)

#logging print commands as output to a file
import sys
log_file = open(base_dir+"/Simulation/output.log", "w")
sys.stdout = log_file


# Importing libraries
import math 
import openpyxl
from sympy import symbols, Eq, solve, log
import pandas as pd
from coolprop import Refr
from Compressor import Compressor
from Condenser import Condenser
from Evaporator import Evaporator
from EXV import EXV
import matplotlib.pyplot as plt
from Plot_OperatingPoints import OP


# Read the refrigeration components excel data
wb_RC = openpyxl.load_workbook(base_dir+"/Inputs/Component_Sheets/Refrigeration_Components.xlsx")
# Compressor data
Compressor_sheet = wb_RC["Compressor"]
Compressor_dict={"Compressor_CC": Compressor_sheet["B2"].value, # compressor displacement
                 "Compressor_efficiency_csv" : Compressor_sheet["B3"].value, # Compressor efficiency table
                 "Compressor_eta_m" : Compressor_sheet["B4"].value} # Compressor mexhanical efficiency
# Condenser data
Condenser_sheet = wb_RC["Condenser"]
Condenser_dict={"H_cond":Condenser_sheet["B3"].value, # condenser height [mm]
                "W_cond":Condenser_sheet["B4"].value, # condenser width [mm] 
                "Depth_cond":Condenser_sheet["B5"].value, # condenser slab depth [mm]
                "N_tube":Condenser_sheet["B7"].value,# total number of tubes  
                "N_passes":Condenser_sheet["B8"].value, # Number of passes 
                "Tubes_split":Condenser_sheet["B9"].value, # Tubes split is different passes
                "Thickness_tube":Condenser_sheet["B10"].value,  # height of tube in mm
                "N_ports": Condenser_sheet["B11"].value, # number of ports in each tube
                "D_char": Condenser_sheet["B12"].value, # height of port triangles [mm]
                "P_fin":Condenser_sheet["B14"].value,  # fin pitch [mm]
                "Thickness_fin":Condenser_sheet["B15"].value,  # thickness of fin [mm]
                "P_louver":Condenser_sheet["B16"].value,  # louver pitch [mm]
                "L_louver":Condenser_sheet["B17"].value, # louver length [mm]
                "Theta_louver":Condenser_sheet["B18"].value} # louver angle [deg]
# Evaporator data
Evaporator_sheet = wb_RC["Evaporator"]
Evaporator_dict={"H_cond":Evaporator_sheet["B3"].value, # condenser height [mm]
                "W_cond":Evaporator_sheet["B4"].value, # condenser width [mm] 
                "Depth_cond":Evaporator_sheet["B5"].value, # condenser slab depth [mm]
                "N_tube":Evaporator_sheet["B7"].value,# total number of tubes  
                "N_passes":Evaporator_sheet["B8"].value, # Number of passes 
                "Tubes_split":Evaporator_sheet["B9"].value, # Tubes split is different passes
                "Thickness_tube":Evaporator_sheet["B10"].value,  # height of tube in mm
                "N_ports": Evaporator_sheet["B11"].value, # number of ports in each tube
                "D_char": Evaporator_sheet["B12"].value, # height of port triangles [mm]
                "P_fin":Evaporator_sheet["B14"].value,  # fin pitch [mm]
                "Thickness_fin":Evaporator_sheet["B15"].value,  # thickness of fin [mm]
                "P_louver":Evaporator_sheet["B16"].value,  # louver pitch [mm]
                "L_louver":Evaporator_sheet["B17"].value, # louver length [mm]
                "Theta_louver":Evaporator_sheet["B18"].value} # louver angle [deg]


# Read boundary conditions input excel data
wb_BC = openpyxl.load_workbook(base_dir+"/Inputs/Boundary_Conditions/Boundary_Conditions.xlsx")
BoundaryConditions_sheet = wb_BC["Boundary_Conditions"]
BoundaryConditions_dict={"Refrigerant":BoundaryConditions_sheet["B2"].value, # Condenser air inlet temperature[°C] 
                        "T_air_cond":BoundaryConditions_sheet["B4"].value, # Condenser air inlet temperature[°C]
                        "P_air_cond":BoundaryConditions_sheet["B5"].value, # Condenser Air inlet Pressure[barA]
                        "RH_air_cond":BoundaryConditions_sheet["B6"].value, # Condenser Air inlet humidity[%]  
                        "M_air_cond":BoundaryConditions_sheet["B7"].value, # Condenser Air mass flow rate [kg/h]                                            
                        "T_air_evap":BoundaryConditions_sheet["B9"].value, # Evaporator air inlet temperature[kg/h]
                        "P_air_evap":BoundaryConditions_sheet["B10"].value, # Evaporator Air inlet Pressure[barA]
                        "RH_air_evap":BoundaryConditions_sheet["B11"].value, # Evaporator Air inlet humidity[%]                          
                        "M_air_evap":BoundaryConditions_sheet["B12"].value, # Evaporator Air mass flow rate [kg/h]
                        "Compressor_RPM":BoundaryConditions_sheet["B14"].value,# Compressor RPM
                        "Input_Subcool":BoundaryConditions_sheet["B16"].value,# Subcolling at condenser outlet
                        "Input_Superheat":BoundaryConditions_sheet["B17"].value}#Superheating at evaporator outlet



# Import Refrigerant diagram
import importlib
module_name = "{}_diag".format(BoundaryConditions_dict["Refrigerant"])
print(module_name,'module_name')
ph_diag_module = importlib.import_module("ph_diagram")
refr_ph_diag = getattr(ph_diag_module, module_name)


# Initial guess for suction pressure is the saturation pressure corresponding to Air inlet temperature at evaporator
Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],3,BoundaryConditions_dict["T_air_evap"],0.9) # Calling function to import refrigerant properties
Ps_max=pressure-1 # initializing the guessed suction pressure list

# Initial guess for discharge pressure is the saturation pressure corresponding to Air inlet temperature at condenser
Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],3,BoundaryConditions_dict["T_air_cond"],0.9) # Calling function to import refrigerant properties
Pd_min=pressure # initializing the guessed discharge pressure list

# Importing R134a ph diagram
(points,line,fig,ax)=refr_ph_diag()
OP(points,line,[0,0,0,0,0],[0,0,0,0,0])


##### Calculation 1 ###### 

Subcool=0
Pd_max=Pd_min+1

# Max number of iterations to avoid infinite loop
max_iterations = 100
 
for iter in range(max_iterations):
    # Refrigerant properties at evaporator outlet
    Refrigerant, Refrigerant_state, pressure, rho, Tempr, enthalpy, entropy, C_p, lambda_K, viscocity_fluid, x, Superheat_Subcool = Refr(BoundaryConditions_dict["Refrigerant"],7, Ps_max, BoundaryConditions_dict["Input_Superheat"])

    # Calling compressor function to estimate refrigerant state at compressor outlet
    P_Compressor_in = pressure
    h_Compressor_in = enthalpy
    m_dot, h_Compressor_out, power_compressor = Compressor(
        Ps=Ps_max,
        Pd=Pd_max,
        rho_Compressor_in=rho,
        h_Compressor_in=enthalpy,
        entropy_Compressor_in=entropy,
        Compressor_dict=Compressor_dict,
        BoundaryConditions_dict=BoundaryConditions_dict
    )

    # Refrigerant properties at compressor outlet
    Refrigerant, Refrigerant_state, pressure, rho, Tempr, enthalpy, entropy, C_p, lambda_K, viscocity_fluid, x, Superheat_Subcool = Refr(BoundaryConditions_dict["Refrigerant"],6, Pd_max, h_Compressor_out * 1000)
    P_Compressor_out = pressure
    h_Compressor_out = enthalpy

    # Condenser outlet state
    P_Cond_out, h_Cond_out, Q_dot_Cond, x_Cond_out, Cond_subcool_out, T_air_Cond_out = Condenser(
        P_Cond_in=pressure,
        T_Cond_in=Tempr,
        m_dot_ref=m_dot,
        Condenser_dict=Condenser_dict,
        BoundaryConditions_dict=BoundaryConditions_dict
    )

    Subcool = Cond_subcool_out


    # Break condition-exit for loop if subcooling is greater than target subcooling
    if Subcool <= BoundaryConditions_dict["Input_Subcool"]:
        break

    # Update Pd_max to help reach target subcooling
    Pd_max += 1 * max(0.1, (1 - Cond_subcool_out / (BoundaryConditions_dict["Input_Subcool"] - 1)))

print('iterations',iter) 


# Regrigerant properties at evaporator inlet 
Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],6,Ps_max, h_Cond_out)
# Calling Evaporator function to estimate refrigerant state at evaporator outlet
P_Evap_out,h_Evap_out,Q_dot_Evap,x_evap_out,Evap_Superheat_out,T_air_Evap_out,RH_air_Evap_out,Q_dot_sensible=Evaporator(P_Evap_in=Ps_max,x_Evap_in=x,m_dot_ref=m_dot,Evaporator_dict=Evaporator_dict,BoundaryConditions_dict=BoundaryConditions_dict) 

# Plottting operating points
OP(points,line,[h_Compressor_in,h_Compressor_out,h_Cond_out/1000,h_Cond_out/1000, h_Evap_out/1000],[Ps_max ,P_Compressor_out, P_Cond_out,Ps_max, P_Evap_out])


print("Calculation 1 results")
print(h_Compressor_in,h_Compressor_out,h_Cond_out/1000,h_Cond_out/1000, h_Evap_out/1000, Ps_max ,P_Compressor_out, P_Cond_out,Ps_max, P_Evap_out)


##### Calculation 2 ###### 

print('Calculation 2 start', Subcool,BoundaryConditions_dict["Input_Subcool"])

Pd_guess=P_Compressor_out-0.2
Ps_guess=Ps_max


Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],7,Ps_guess,BoundaryConditions_dict["Input_Superheat"]) # Calling function to import refrigerant properties
h_Evap_out_target=enthalpy*1000



# Max number of iterations to avoid infinite loop
max_iterations_outer = 100
max_iterations_inner = 100



for outer_iter in range(max_iterations_outer):  
    for inner_iter in range(max_iterations_inner):
        # Refrigerant properties at evaporator outlet
        Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],7,Ps_guess,BoundaryConditions_dict["Input_Superheat"])
        

        # Calling compressor function to estimate refrigerant state at compressor outlet
        P_Compressor_in=pressure
        h_Compressor_in=enthalpy
        m_dot,h_Compressor_out,power_compressor=Compressor(Ps=Ps_guess,Pd=Pd_guess,rho_Compressor_in=rho, h_Compressor_in=enthalpy, entropy_Compressor_in=entropy,Compressor_dict=Compressor_dict,BoundaryConditions_dict=BoundaryConditions_dict)
        Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],6,Pd_guess,h_Compressor_out*1000)  # Obtaining refrigerant properties at compressor outlet
        P_Compressor_out=pressure
        h_Compressor_out=enthalpy

        # Calling condenser function to estimate refrigerant state at condenser outlet
        P_Cond_out,h_Cond_out,Q_dot_Cond,x_Cond_out,Cond_subcool_out,T_air_Cond_out=Condenser(P_Cond_in=pressure,T_Cond_in=Tempr,m_dot_ref =m_dot,Condenser_dict=Condenser_dict,BoundaryConditions_dict=BoundaryConditions_dict) 
        Subcool=Cond_subcool_out

        # Exit inner loop if subcooling target is met
        if Subcool <= BoundaryConditions_dict["Input_Subcool"]:
            break

        Ps_guess=Ps_guess-max(0.02,0.4*(1-Cond_subcool_out/(BoundaryConditions_dict["Input_Subcool"]-1)))
        
    
    # Regrigerant properties at evaporator inlet 
    Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],6,P_Compressor_in, h_Cond_out)
    P_Evap_in=pressure

    # Calling Evaporator function to estimate refrigerant state at evaporator outlet
    P_Evap_out,h_Evap_out,Q_dot_Evap,x_evap_out,Evap_Superheat_out,T_air_Evap_out,RH_air_Evap_out,Q_dot_sensible=Evaporator(P_Evap_in=P_Evap_in,x_Evap_in=x,m_dot_ref=m_dot,Evaporator_dict=Evaporator_dict,BoundaryConditions_dict=BoundaryConditions_dict) 

    Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],7,P_Evap_in,BoundaryConditions_dict["Input_Superheat"]) # Calling function to import refrigerant properties
    h_Evap_out_target=enthalpy*1000


    OP(points,line,[h_Compressor_in,h_Compressor_out,h_Cond_out/1000,h_Cond_out/1000, h_Evap_out/1000],[P_Compressor_in ,P_Compressor_out, P_Cond_out,P_Evap_in, P_Evap_out])

    # Exit outer loop if evaporator outlet enthalpy target is met
    if h_Evap_out >= h_Evap_out_target:
        break

    Pd_guess=Pd_guess-max(0.1,2*(1-h_Evap_out/h_Evap_out_target))



# Plottting operating points
OP(points,line,[h_Compressor_in,h_Compressor_out,h_Cond_out/1000,h_Cond_out/1000, h_Evap_out/1000],[P_Compressor_in  ,P_Compressor_out, P_Cond_out,P_Evap_in, P_Evap_out])



##### Calculation 3 ###### 

print('Calculation 3 start', Subcool,BoundaryConditions_dict["Input_Subcool"])

Pd_guess=P_Compressor_out
Ps_guess=P_Compressor_in
P_Evap_in=P_Evap_in+(P_Compressor_in-P_Evap_out)/2



Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],7,Ps_guess,BoundaryConditions_dict["Input_Superheat"]) # Calling function to import refrigerant properties
h_Evap_out_target=enthalpy*1000


# Max number of iterations to avoid infinite loop
max_iterations_outer = 100
max_iterations_inner = 100
max_iterations_utmost = 100


for utmost_iter in range(max_iterations_utmost):  

    for outer_iter in range(max_iterations_outer):      
        for inner_iter in range(max_iterations_inner):
            # Refrigerant properties at evaporator outlet
            Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],7,Ps_guess,BoundaryConditions_dict["Input_Superheat"])
            
            # Calling compressor function to estimate refrigerant state at compressor outlet
            P_Compressor_in=pressure
            h_Compressor_in=enthalpy

            m_dot,h_Compressor_out,power_compressor=Compressor(Ps=Ps_guess,Pd=Pd_guess,rho_Compressor_in=rho, h_Compressor_in=enthalpy, entropy_Compressor_in=entropy,Compressor_dict=Compressor_dict,BoundaryConditions_dict=BoundaryConditions_dict)
            Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],6,Pd_guess,h_Compressor_out*1000)  # Obtaining refrigerant properties at compressor outlet
            P_Compressor_out=pressure
            h_Compressor_out=enthalpy

            # Calling condenser function to estimate refrigerant state at condenser outlet
            P_Cond_out,h_Cond_out,Q_dot_Cond,x_Cond_out,Cond_subcool_out,T_air_Cond_out=Condenser(P_Cond_in=pressure,T_Cond_in=Tempr,m_dot_ref =m_dot,Condenser_dict=Condenser_dict,BoundaryConditions_dict=BoundaryConditions_dict) 
            Subcool=Cond_subcool_out


            # Exit inner loop if subcooling target is met
            if abs(Subcool-BoundaryConditions_dict["Input_Subcool"]) <1:
                break

            Ps_guess=Ps_guess-0.2*(1-Cond_subcool_out/(BoundaryConditions_dict["Input_Subcool"]-1))
        
        
        # Regrigerant properties at evaporator inlet 
        Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],6,P_Evap_in, h_Cond_out)
        # Calling Evaporator function to estimate refrigerant state at evaporator outlet
        P_Evap_out,h_Evap_out,Q_dot_Evap,x_evap_out,Evap_Superheat_out,T_air_Evap_out,RH_air_Evap_out,Q_dot_sensible=Evaporator(P_Evap_in=P_Evap_in,x_Evap_in=x,m_dot_ref=m_dot,Evaporator_dict=Evaporator_dict,BoundaryConditions_dict=BoundaryConditions_dict) 

        Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],7,P_Evap_out,BoundaryConditions_dict["Input_Superheat"]) # Calling function to import refrigerant properties
        h_Evap_out_target=enthalpy*1000

        OP(points,line,[h_Compressor_in,h_Compressor_out,h_Cond_out/1000,h_Cond_out/1000, h_Evap_out/1000],[P_Compressor_in ,P_Compressor_out, P_Cond_out,P_Evap_in, P_Evap_out])

        # Exit outer loop if evaporator outlet enthalpy target is met
        if abs(h_Evap_out-h_Evap_out_target)<1000:
            break

        Pd_guess=Pd_guess-0.5*(1-h_Evap_out/h_Evap_out_target)

    OP(points,line,[h_Compressor_in,h_Compressor_out,h_Cond_out/1000,h_Cond_out/1000, h_Evap_out/1000],[P_Compressor_in ,P_Compressor_out, P_Cond_out,P_Evap_in, P_Evap_out])

        # Exit outer loop if evaporator outlet  and compressor inlet pressure is met
    if abs(P_Evap_out-P_Compressor_in)<0.1:
        break
    Pd_guess=P_Compressor_out
    Ps_guess=P_Compressor_in
    P_Evap_in=P_Evap_in+(P_Compressor_in-P_Evap_out)/2
   
   
   
# Plottting operating points
OP(points,line,[h_Compressor_in,h_Compressor_out,h_Cond_out/1000,h_Cond_out/1000, h_Evap_out/1000],[P_Compressor_in  ,P_Compressor_out, P_Cond_out,P_Evap_in, P_Evap_out])



Refrigerant,Refrigerant_state,pressure,rho,Tempr,enthalpy,entropy, C_p, lambda_K,viscocity_fluid,x,Superheat_Subcool=Refr(BoundaryConditions_dict["Refrigerant"],6,P_Cond_out,h_Cond_out)
Orifice_area,Orifice_dia=EXV(m_dot_ref=m_dot,Ps=P_Evap_in,Pd=P_Cond_out,rho_EXV_in=rho) 


plt.savefig(base_dir+"/Simulation/Final_ph_diagram.jpeg", format="jpeg")


end_time = time.time()
execution_time = end_time - start_time
print("Execution time:", execution_time, "seconds")


os.chdir(base_dir+'/Simulation')

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import shutil

# Specify the source file and destination path
source_file = base_dir+"/Templates/Results_Template.xlsx"
from datetime import datetime
current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
destination_file = base_dir+f"/Results_Record/Result_{current_time}.xlsx"


# Copy the file and rename it
shutil.copy(source_file, destination_file)


# Write the output sheet
wb = load_workbook(destination_file)
ws = wb.active  
ws["B2"]=h_Compressor_in
ws["B3"]=h_Compressor_out
ws["B4"]=h_Cond_out/1000
ws["B5"]=h_Cond_out/1000
ws["B6"]=h_Evap_out/1000
ws["C2"]=P_Compressor_in
ws["C3"]=P_Compressor_out
ws["C4"]=P_Cond_out
ws["C5"]=P_Evap_in
ws["C6"]=P_Evap_out
ws["B9"]=m_dot*3600
ws["B10"]=power_compressor*1000
ws["B13"]=Q_dot_Evap*-1
ws["B14"]=Q_dot_sensible
ws["B15"]=T_air_Evap_out
ws["B16"]=RH_air_Evap_out*100
ws["B19"]=Q_dot_Cond 
ws["B20"]=T_air_Cond_out
ws["B23"]=Orifice_area
ws["B24"]=Orifice_dia




# Create an Image object and add it to the worksheet
img = Image(base_dir+"/Simulation/Final_ph_diagram.jpeg")
ws.add_image(img, "I2")  # Specify the cell where the image will be placed

# Save the workbook
wb.save(destination_file)

log_file.close()

os.chdir(base_dir)









    





 